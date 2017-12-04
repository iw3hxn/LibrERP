# -*- encoding: utf-8 -*-
##############################################################################
#
# Copyright (c) 2015 Andrei Levin (andrei.levin at didotech.com)
#
#                          All Rights Reserved.
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
###############################################################################

import logging

from openerp.osv import orm
from tools.translate import _

_logger = logging.getLogger(__name__)


def import_sheet(filename, content):
    name, file_type = filename.rsplit('.', 1)

    if file_type in ('xls', 'xlsb'):
        # "Excel"
        try:
            import xlrd
        except ImportError:
            _logger.error('Cannot `import xlrd`.')

        for encoding in ('utf-8', 'latin-1', 'cp1252'):
            try:
                book = xlrd.open_workbook(file_contents=content, encoding_override=encoding)
                break
            except UnicodeDecodeError:
                pass
        else:
            raise orm.except_orm('Error', _('Unknown encoding'))

        table = []
        sh = book.sheet_by_index(0)

        for rx in range(sh.nrows):
            row = []
            for cx in range(sh.ncols):
                row.append(sh.cell(rowx=rx, colx=cx).value)
            table.append(row)
#            print rx
#            print table
        number_of_lines = sh.nrows
    elif file_type in ('xlsx', 'xlsm', 'xltx', 'xltm'):
        try:
            import openpyxl
        except ImportError:
            _logger.error('Cannot `import openpyxl`.')
        import StringIO

        ## Create virtual File:
        virtual_file = StringIO.StringIO(content)
        book = openpyxl.load_workbook(virtual_file)

        table = []
        sh = book.worksheets[0]
        max_column = sh.max_column

        for rx in range(1, sh.max_row + 1):
            row = []
            for cx in range(1, max_column + 1):
                if rx == 1 and not sh.cell(row=rx, column=cx).value:
                    max_column = cx - 1
                    break

                row.append(sh.cell(row=rx, column=cx).value)
            table.append(row)
        number_of_lines = sh.max_row
    elif file_type in ('ods', ):
        # "OpenOffice"
        try:
            from openerp.addons.core_extended.odf_to_array import ODSReader
        except ImportError:
            _logger.error('Cannot `import OpenOffice`.')
        import StringIO

        ## Create virtual File:
        virtual_file = StringIO.StringIO(content)

        book = ODSReader(virtual_file)
        table = book.sheet_by_index(0)
        number_of_lines = len(table)
    elif file_type == 'csv':
        # "CSV"
        import csv
        import StringIO

        ## Create virtual File:
        virtual_file = StringIO.StringIO(content)

        ## Process CSV file:
        sample = virtual_file.read(512)
        virtual_file.seek(0)

        dialect = csv.Sniffer().sniff(sample)
        table = csv.reader(virtual_file, dialect)

        # self.table is an object of type '_csv.reader' and has no len() method
        number_of_lines = sum(1 for row in table)
        virtual_file.seek(0)
    else:
        raise orm.except_orm(_('Error'), _('Unknown file extension'))

    return table, number_of_lines
